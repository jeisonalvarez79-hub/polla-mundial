"""
Genera un Excel que documenta el sistema de puntuacion completo usado en
ambas pollas (valores tomados de config.pts en Supabase + las constantes de
bracket/bonus de src/data/initialData.js, que no son configurables desde Admin).
Ejecutar: python scripts/build-sistema-puntuacion-excel.py <output.xlsx>
"""
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='2F6B4D')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
SECTION_FILL = PatternFill('solid', fgColor='EEF2EA')
SECTION_FONT = Font(bold=True, size=11, color='2F6B4D')
GOLD_FILL = PatternFill('solid', fgColor='FBF1DE')
TOTAL_FONT = Font(bold=True, size=10)
THIN = Side(style='thin', color='CCD4C3')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# Valores reales tomados de Supabase (config.pts) el 2026-07-21
PTS_GRUPOS = {'exacto': 3, 'resultado': 1, 'ordenGrupo': 1, 'clasificado': 2, 'goleador': 10}
# Constantes de bracket/bonus (src/data/initialData.js) - no configurables desde Admin
BRACKET_PAIRING_PTS = {'r32': 3, 'r16': 4, 'qf': 8, 'sf': 8, 'third': 8, 'final': 12}
BRACKET_TEAM_PTS = {'r16': 2, 'qf': 4, 'sf': 4, 'finalFour': 8}
BONUS_PTS = {'champion': 15, 'runnerUp': 10, 'thirdPlace': 8, 'fourthPlace': 6}


def header_row(ws, values, widths=None):
    ws.append(values)
    for c in ws[ws.max_row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if widths:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w


def section_row(ws, text, span=4):
    ws.append([text] + [''] * (span - 1))
    row = ws.max_row
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1)
    c.fill = SECTION_FILL
    c.font = SECTION_FONT
    c.alignment = Alignment(horizontal='left', vertical='center')


def data_row(ws, values, bold_last=False, fill=None):
    ws.append(values)
    row = ws.max_row
    for c in ws[row]:
        c.border = BORDER
    if bold_last:
        ws.cell(row=row, column=len(values)).font = TOTAL_FONT
    if fill:
        for c in ws[row]:
            c.fill = fill


def autosize(ws, min_width=8, max_width=60):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def main():
    output_path = sys.argv[1]
    wb = Workbook()

    # ══════════════════ Sheet 1: Resumen ══════════════════
    ws = wb.active
    ws.title = 'Resumen'
    header_row(ws, ['Categoría', 'Cómo se gana', 'Puntos', 'Máximo posible'], widths=[26, 55, 10, 16])

    section_row(ws, 'FASE DE GRUPOS')
    data_row(ws, ['Marcador exacto', 'Acertar el marcador exacto de un partido de grupos', PTS_GRUPOS['exacto'], '216 (72 partidos x 3)'])
    data_row(ws, ['Resultado correcto', 'Acertar ganador/empate sin el marcador exacto', PTS_GRUPOS['resultado'], 'No se suma al total (excluyente con "Marcador exacto")'])
    data_row(ws, ['Posición exacta en grupo', 'Acertar el equipo que quedó en cada puesto 1°-4° de su grupo', PTS_GRUPOS['ordenGrupo'], '48 (12 grupos x 4 puestos)'])
    data_row(ws, ['Clasificado a 32avos', 'El equipo pronosticado sí clasificó a dieciseisavos (24 líderes/subcampeones + 8 mejores 3eros), sin importar el puesto exacto', PTS_GRUPOS['clasificado'], '64 (32 equipos x 2)'])

    section_row(ws, 'FASE ELIMINATORIA (BRACKET) — equipo clasificado')
    data_row(ws, ['16avos (R32) — equipo', 'No aplica: en esta ronda solo se puntúa la llave, no equipos individuales', '—', '0'])
    data_row(ws, ['Octavos (R16) — equipo', 'Por cada equipo pronosticado que sí llegó a Octavos', BRACKET_TEAM_PTS['r16'], '32 (16 equipos x 2)'])
    data_row(ws, ['Cuartos (QF) — equipo', 'Por cada equipo pronosticado que sí llegó a Cuartos', BRACKET_TEAM_PTS['qf'], '32 (8 equipos x 4)'])
    data_row(ws, ['Semifinal (SF) — equipo', 'Por cada equipo pronosticado que sí llegó a Semifinal', BRACKET_TEAM_PTS['sf'], '16 (4 equipos x 4)'])
    data_row(ws, ['Final / 3er puesto — equipo', 'Por cada equipo pronosticado en el partido correcto (Final o 3er puesto, no son intercambiables)', BRACKET_TEAM_PTS['finalFour'], '32 (4 equipos x 8)'])

    section_row(ws, 'FASE ELIMINATORIA (BRACKET) — llave acertada (ambos equipos, sin importar orden)')
    data_row(ws, ['16avos (R32) — llave', 'Acertar los 2 equipos de una llave de dieciseisavos', BRACKET_PAIRING_PTS['r32'], '48 (16 llaves x 3)'])
    data_row(ws, ['Octavos (R16) — llave', 'Acertar los 2 equipos de una llave de octavos', BRACKET_PAIRING_PTS['r16'], '32 (8 llaves x 4)'])
    data_row(ws, ['Cuartos (QF) — llave', 'Acertar los 2 equipos de una llave de cuartos', BRACKET_PAIRING_PTS['qf'], '32 (4 llaves x 8)'])
    data_row(ws, ['Semifinal (SF) — llave', 'Acertar los 2 equipos de una llave de semifinal', BRACKET_PAIRING_PTS['sf'], '16 (2 llaves x 8)'])
    data_row(ws, ['3er puesto — llave', 'Acertar los 2 equipos del partido de 3er/4to puesto', BRACKET_PAIRING_PTS['third'], '8'])
    data_row(ws, ['Final — llave', 'Acertar los 2 equipos de la Final', BRACKET_PAIRING_PTS['final'], '12'])

    section_row(ws, 'BONUS — PUESTOS FINALES')
    data_row(ws, ['Campeón', 'Acertar el equipo Campeón del Mundial', BONUS_PTS['champion'], '15'])
    data_row(ws, ['Subcampeón', 'Acertar el equipo Subcampeón (perdedor de la Final)', BONUS_PTS['runnerUp'], '10'])
    data_row(ws, ['Tercer puesto', 'Acertar el equipo que gana el partido de 3er/4to puesto', BONUS_PTS['thirdPlace'], '8'])
    data_row(ws, ['Cuarto puesto', 'Acertar el equipo que pierde el partido de 3er/4to puesto', BONUS_PTS['fourthPlace'], '6'])

    section_row(ws, 'GOLEADOR')
    data_row(ws, ['Goleador', 'El goleador pronosticado (1 nombre) aparece en el listado oficial de máximos goleadores', PTS_GRUPOS['goleador'], '10'])

    # No se suma el máximo de "Resultado correcto" (72): es mutuamente excluyente con
    # "Marcador exacto" en un mismo partido (o se gana uno o el otro, nunca ambos).
    total_max = 216 + 48 + 64 + 32 + 32 + 16 + 32 + 48 + 32 + 32 + 16 + 8 + 12 + 15 + 10 + 8 + 6 + 10
    section_row(ws, 'PUNTAJE MÁXIMO TEÓRICO TOTAL')
    data_row(ws, ['TOTAL', 'Suma de todos los máximos de arriba, usando marcador exacto (no resultado, que es excluyente con este) en todos los partidos + acierto perfecto en toda la fase eliminatoria + goleador', '', total_max], bold_last=True, fill=GOLD_FILL)

    ws.freeze_panes = 'A2'
    autosize(ws, min_width=10, max_width=70)
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 62

    # ══════════════════ Sheet 2: Fase de Grupos (detalle) ══════════════════
    ws2 = wb.create_sheet('Fase de Grupos')
    header_row(ws2, ['Concepto', 'Detalle', 'Puntos'], widths=[28, 70, 10])
    data_row(ws2, ['Marcador exacto', 'Ej: pronosticaste 2-1 y el resultado real fue 2-1 → ganas el máximo por ese partido.', PTS_GRUPOS['exacto']])
    data_row(ws2, ['Resultado correcto', 'Ej: pronosticaste 2-1 (gana local) y el resultado real fue 3-0 (también gana local, pero marcador distinto) → ganas este puntaje menor. Si el marcador exacto no aplicó, se evalúa este.', PTS_GRUPOS['resultado']])
    data_row(ws2, ['Fallo', 'Si ni el marcador ni el resultado (ganador/empate) coinciden, el partido no suma puntos.', 0])
    data_row(ws2, ['Posición exacta en grupo', 'Se compara la tabla de posiciones final de cada grupo (criterio FIFA: Pts → Dif. de gol → Goles a favor → Head-to-head) contra la tabla que resulta de tus pronósticos de marcador. Se gana 1 pt por cada puesto (1° a 4°) que coincida exactamente.', PTS_GRUPOS['ordenGrupo']])
    data_row(ws2, ['Clasificado a 32avos', 'Se comparan los 32 equipos reales que avanzaron a dieciseisavos (24 líderes/subcampeones de los 12 grupos + los 8 mejores terceros) contra los 32 equipos que resultan de tus pronósticos de grupo, sin importar en qué posición haya quedado cada uno.', PTS_GRUPOS['clasificado']])
    autosize(ws2, min_width=14, max_width=80)
    ws2.freeze_panes = 'A2'

    # ══════════════════ Sheet 3: Fase Eliminatoria (detalle) ══════════════════
    ws3 = wb.create_sheet('Fase Eliminatoria')
    header_row(ws3, ['Ronda', 'Pts por equipo clasificado', 'Pts por llave acertada (ambos equipos)', 'Notas'], widths=[16, 14, 16, 75])
    data_row(ws3, ['16avos (R32)', '—', BRACKET_PAIRING_PTS['r32'], 'En esta ronda no se puntúa por equipo individual, solo por acertar la llave completa (los 2 equipos que se enfrentan).'])
    data_row(ws3, ['Octavos (R16)', BRACKET_TEAM_PTS['r16'], BRACKET_PAIRING_PTS['r16'], 'Los puntos por equipo son acumulativos: si un equipo avanza varias rondas, suma en cada ronda que alcanzó según tu pronóstico.'])
    data_row(ws3, ['Cuartos (QF)', BRACKET_TEAM_PTS['qf'], BRACKET_PAIRING_PTS['qf'], ''])
    data_row(ws3, ['Semifinal (SF)', BRACKET_TEAM_PTS['sf'], BRACKET_PAIRING_PTS['sf'], ''])
    data_row(ws3, ['Final / 3er puesto', BRACKET_TEAM_PTS['finalFour'], f"{BRACKET_PAIRING_PTS['third']} (3er puesto) / {BRACKET_PAIRING_PTS['final']} (Final)", 'IMPORTANTE: el equipo debe estar en el partido correcto. Si pronosticaste que un equipo llega a la Final pero en realidad perdió la semifinal y jugó el partido de 3er puesto, NO se cuenta como acierto — son roles distintos, no intercambiables.'])
    autosize(ws3, min_width=14, max_width=80)
    ws3.freeze_panes = 'A2'

    # ══════════════════ Sheet 4: Bonus y Goleador ══════════════════
    ws4 = wb.create_sheet('Bonus y Goleador')
    header_row(ws4, ['Concepto', 'Detalle', 'Puntos'], widths=[20, 65, 10])
    section_row(ws4, 'BONUS DE PUESTOS FINALES (se suman aparte de los puntos por equipo/llave)')
    data_row(ws4, ['Campeón', 'Acertar quién sale Campeón del Mundial', BONUS_PTS['champion']])
    data_row(ws4, ['Subcampeón', 'Acertar quién sale Subcampeón (perdedor de la Final)', BONUS_PTS['runnerUp']])
    data_row(ws4, ['Tercer puesto', 'Acertar quién gana el partido por el 3er puesto', BONUS_PTS['thirdPlace']])
    data_row(ws4, ['Cuarto puesto', 'Acertar quién pierde el partido por el 3er puesto (sale 4to)', BONUS_PTS['fourthPlace']])
    section_row(ws4, 'GOLEADOR')
    data_row(ws4, ['Goleador', 'Se pronostica 1 solo jugador. Si ese jugador queda entre los máximos goleadores oficiales del torneo, se gana el puntaje completo (no hay puntaje parcial).', PTS_GRUPOS['goleador']])
    autosize(ws4, min_width=14, max_width=75)
    ws4.freeze_panes = 'A2'

    wb.save(output_path)
    print('OK ->', output_path)


if __name__ == '__main__':
    main()
