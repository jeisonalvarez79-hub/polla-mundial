"""
Genera una "Polla Mundial" en Excel para llenar 100% a mano: un participante
llena su pronostico, otra persona llena los resultados reales, y las formulas
asignan los puntos automaticamente hasta el campeon, incluyendo clasificados
de grupos, llaves de bracket y goleador.

Version SIMPLE (acordada con el usuario): el desempate de tabla de grupos usa
Puntos -> Diferencia de gol -> Goles a favor (sin el mini-cuadrangular de
enfrentamientos directos que usa la app real), y el emparejamiento del
bracket se llena a mano en cada ronda (no se autoencadena desde los
pronosticos de grupo).

Ejecutar: python scripts/build-manual-template-excel.py <schedule.json> <output.xlsx>
"""
import json
import re
import sys
from collections import OrderedDict

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.workbook.properties import CalcProperties

HEADER_FILL = PatternFill('solid', fgColor='2F6B4D')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
SECTION_FILL = PatternFill('solid', fgColor='EEF2EA')
SECTION_FONT = Font(bold=True, size=11, color='2F6B4D')
INPUT_FILL = PatternFill('solid', fgColor='FFFCE8')
CALC_FILL = PatternFill('solid', fgColor='F4F6F0')
GOLD_FILL = PatternFill('solid', fgColor='FBF1DE')
THIN = Side(style='thin', color='CCD4C3')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROUND_LABELS = {'r32': '32avos', 'r16': '16avos', 'qf': 'Cuartos', 'sf': 'Semifinal', 'third': '3er Puesto', 'final': 'Final'}
ROUND_ORDER = ['r32', 'r16', 'qf', 'sf', 'third', 'final']
ROUND_PAIR_PTS = {'r32': 3, 'r16': 4, 'qf': 8, 'sf': 8, 'third': 8, 'final': 12}
ROUND_TEAM_PTS = {'r16': 2, 'qf': 4, 'sf': 4}

GROUP_LETTERS = list('ABCDEFGHIJKL')


def colletter(idx):
    return get_column_letter(idx)


def membership_formula(real_cells, pred_range, pts):
    """
    Puntos por 'equipo clasificado': por cada celda real en real_cells que SI
    aparece en algun lugar de pred_range, suma pts. Usa COUNTIF con criterio
    escalar (una celda a la vez) en vez de COUNTIF(rango,rango) dentro de
    SUMPRODUCT — ese patron de 'criterio como rango' no se evalua de forma
    confiable en todos los motores de calculo, aunque en Excel de escritorio
    normalmente funciona. Esta forma es mas verbosa pero 100% universal.
    """
    terms = [f"MIN(COUNTIF({pred_range},{c}),1)" for c in real_cells]
    return f"=({'+'.join(terms)})*{pts}"


def style_header(cell):
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def style_section(ws, row, text, span):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1)
    c.fill = SECTION_FILL
    c.font = SECTION_FONT


def autosize(ws, min_width=9, max_width=42):
    widths = {}
    for row in ws.iter_rows():
        for c in row:
            if c.value is None:
                continue
            col = c.column
            length = len(str(c.value))
            widths[col] = max(widths.get(col, 0), length)
    for col, length in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = max(min_width, min(max_width, length + 2))


def main():
    schedule_path, output_path = sys.argv[1], sys.argv[2]
    with open(schedule_path, encoding='utf-8') as f:
        sched = json.load(f)

    matches = sched['matches']
    matches.sort(key=lambda m: (m['group'], int(re.sub(r'\D', '', m['id']))))
    bracket = sched['bracketMatches']

    groups_teams = OrderedDict()
    for m in matches:
        groups_teams.setdefault(m['group'], [])
        for t in (m['homeTeam'], m['awayTeam']):
            if t not in groups_teams[m['group']]:
                groups_teams[m['group']].append(t)

    wb = Workbook()
    wb.calculation = CalcProperties(fullCalcOnLoad=True)

    # ══════════════════════════ 1. INSTRUCCIONES ══════════════════════════
    ws = wb.active
    ws.title = 'Instrucciones'
    ws.column_dimensions['A'].width = 100
    lines = [
        ('Polla Mundial 2026 — Formato manual en Excel', True, 16),
        ('', False, 10),
        ('Como usar este archivo:', True, 12),
        ('1. Hoja "Grupos": una persona (el participante) llena las columnas "Pron. Local" y "Pron. Visitante" con su '
         'pronostico de marcador para cada uno de los 72 partidos. Otra persona (el administrador) llena "Real Local" '
         'y "Real Visitante" a medida que se juegan los partidos reales.', False, 10),
        ('2. Hoja "Tabla Grupos": se calcula sola (no se toca) — arma la tabla de posiciones de cada grupo, tanto real '
         'como segun el pronostico, y compara quien clasifico a Dieciseisavos (32 equipos).', False, 10),
        ('3. Hoja "Bracket": para cada llave (32avos a Final), el participante escribe manualmente que 2 equipos cree '
         'que se enfrentan y quien gana; el administrador llena los equipos y el ganador real. Los puntos de equipo '
         'clasificado y de llave acertada se calculan solos.', False, 10),
        ('4. Hoja "Goleador": el participante escribe 1 nombre; el administrador escribe hasta 3 nombres reales '
         '(estos son los "maximos goleadores" oficiales del torneo).', False, 10),
        ('5. Hoja "Resumen Puntaje": muestra el total, fase por fase, calculado automaticamente. No se toca.', False, 10),
        ('', False, 10),
        ('IMPORTANTE — simplificaciones de esta version (frente al motor real de la app):', True, 11),
        ('- El desempate de la tabla de grupos usa Puntos -> Diferencia de gol -> Goles a favor. NO incluye el '
         'mini-cuadrangular de enfrentamientos directos (head-to-head) que usa la FIFA cuando 2+ equipos quedan '
         'exactamente empatados en esos 3 criterios. En ese caso, el desempate aqui es arbitrario (por el orden en '
         'que aparecen los equipos).', False, 10),
        ('- El bracket NO se arma solo a partir de tus pronosticos de grupo: cada ronda se llena a mano, incluyendo '
         'que equipos crees que se enfrentan.', False, 10),
        ('- Las llaves se comparan por posicion (la llave #5 real contra la llave #5 pronosticada), no por "si esa '
         'pareja de equipos ocurrio en cualquier posicion de la ronda" como hace la app.', False, 10),
        ('- El nombre del goleador se compara sin distinguir mayusculas/minusculas ni espacios extra, pero SI debe '
         'ser el mismo nombre (ej: "Mbappe" y "Kylian Mbappe" cuentan distinto salvo que escribas exactamente igual '
         'en ambos lados).', False, 10),
        ('', False, 10),
        ('Puntajes usados (iguales a los de la app):', True, 11),
        ('Grupos: marcador exacto 3 pts | resultado correcto 1 pt | posicion exacta en tabla 1 pt/puesto | '
         'clasificado a 32avos 2 pts/equipo', False, 10),
        ('Bracket — equipo clasificado: Octavos 2 pts | Cuartos 4 pts | Semifinal 4 pts | Final/3er puesto 8 pts', False, 10),
        ('Bracket — llave acertada: 32avos 3 pts | Octavos 4 pts | Cuartos 8 pts | Semifinal 8 pts | 3er puesto 8 pts | Final 12 pts', False, 10),
        ('Bonus: Campeon 15 pts | Subcampeon 10 pts | 3er puesto 8 pts | 4to puesto 6 pts', False, 10),
        ('Goleador: 10 pts (todo o nada)', False, 10),
    ]
    r = 1
    for text, bold, size in lines:
        c = ws.cell(row=r, column=1, value=text)
        c.font = Font(bold=bold, size=size, color='2F6B4D' if bold else '17231A')
        c.alignment = Alignment(wrap_text=True, vertical='top')
        r += 1

    # ══════════════════════════ 2. EQUIPOS (referencia) ══════════════════════════
    ws_eq = wb.create_sheet('Equipos')
    ws_eq.append(['Grupo', 'Equipo'])
    for c in ws_eq[1]:
        style_header(c)
    for g, teams in groups_teams.items():
        for t in teams:
            ws_eq.append([g, t])
    autosize(ws_eq)
    ws_eq.freeze_panes = 'A2'

    # ══════════════════════════ 3. GRUPOS ══════════════════════════
    ws_g = wb.create_sheet('Grupos')
    headers = ['#', 'Grupo', 'Local', 'Visitante', 'Pron. Local', 'Pron. Visitante', 'Real Local', 'Real Visitante', 'Puntos']
    ws_g.append(headers)
    for c in ws_g[1]:
        style_header(c)
    GROUPS_FIRST_ROW = 2
    row = GROUPS_FIRST_ROW
    for i, m in enumerate(matches, start=1):
        ws_g.append([i, m['group'], m['homeTeam'], m['awayTeam'], None, None, None, None, None])
        r_ = ws_g.max_row
        f = f"=IF(OR(G{r_}=\"\",H{r_}=\"\",E{r_}=\"\",F{r_}=\"\"),0,IF(AND(E{r_}=G{r_},F{r_}=H{r_}),3,IF(SIGN(E{r_}-F{r_})=SIGN(G{r_}-H{r_}),1,0)))"
        ws_g.cell(row=r_, column=9, value=f)
        for col in (5, 6, 7, 8):
            ws_g.cell(row=r_, column=col).fill = INPUT_FILL
        ws_g.cell(row=r_, column=9).fill = CALC_FILL
        for c in ws_g[r_]:
            c.border = BORDER
    GROUPS_LAST_ROW = ws_g.max_row
    ws_g.append(['', '', '', '', '', '', '', 'TOTAL GRUPOS', f"=SUM(I{GROUPS_FIRST_ROW}:I{GROUPS_LAST_ROW})"])
    for c in ws_g[ws_g.max_row]:
        c.font = Font(bold=True)
        c.fill = GOLD_FILL
    autosize(ws_g)
    ws_g.freeze_panes = 'A2'
    ws_g.column_dimensions['C'].width = 18
    ws_g.column_dimensions['D'].width = 18

    GR = f"Grupos!$B${GROUPS_FIRST_ROW}:$B${GROUPS_LAST_ROW}"   # Grupo column
    GC_ = f"Grupos!$C${GROUPS_FIRST_ROW}:$C${GROUPS_LAST_ROW}"  # Local
    GD = f"Grupos!$D${GROUPS_FIRST_ROW}:$D${GROUPS_LAST_ROW}"   # Visitante
    GE = f"Grupos!$E${GROUPS_FIRST_ROW}:$E${GROUPS_LAST_ROW}"   # Pron Local
    GF_ = f"Grupos!$F${GROUPS_FIRST_ROW}:$F${GROUPS_LAST_ROW}"  # Pron Visitante
    GG = f"Grupos!$G${GROUPS_FIRST_ROW}:$G${GROUPS_LAST_ROW}"   # Real Local
    GH = f"Grupos!$H${GROUPS_FIRST_ROW}:$H${GROUPS_LAST_ROW}"   # Real Visitante

    # ══════════════════════════ 4. TABLA GRUPOS ══════════════════════════
    ws_t = wb.create_sheet('Tabla Grupos')
    team_headers = [
        'Grupo', 'Equipo',
        'J', 'G', 'E', 'P', 'GF', 'GC', 'DG', 'Pts', 'OrdenReal', 'PosReal',
        'J', 'G', 'E', 'P', 'GF', 'GC', 'DG', 'Pts', 'OrdenPron', 'PosPron',
        'Posicion', 'EquipoReal(pos)', 'EquipoPron(pos)', 'PtsPosicion',
        'PtsReal(pos)', 'DGReal(pos)', 'GFReal(pos)',
        'PtsPron(pos)', 'DGPron(pos)', 'GFPron(pos)',
    ]
    ws_t.append(team_headers)
    for c in ws_t[1]:
        style_header(c)
    style_section(ws_t, 1, '', 1)  # no-op placeholder to keep structure simple

    TEAM_FIRST_ROW = 2
    group_row_ranges = {}
    row = TEAM_FIRST_ROW
    for gi, g in enumerate(GROUP_LETTERS):
        teams = groups_teams[g]
        r0 = row
        for ti, team in enumerate(teams):
            rr = row
            ws_t.cell(row=rr, column=1, value=g)
            ws_t.cell(row=rr, column=2, value=team)
            # Real stats (C..J = 3..10)
            ws_t.cell(row=rr, column=3, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*(({GC_}=$B{rr})+({GD}=$B{rr}))*({GG}<>\"\"))"))
            ws_t.cell(row=rr, column=4, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*({GC_}=$B{rr})*({GG}>{GH}))+SUMPRODUCT(({GR}=$A{rr})*({GD}=$B{rr})*({GH}>{GG}))"))
            ws_t.cell(row=rr, column=5, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*(({GC_}=$B{rr})+({GD}=$B{rr}))*({GG}={GH})*({GG}<>\"\"))"))
            ws_t.cell(row=rr, column=6, value=f"=C{rr}-D{rr}-E{rr}")
            ws_t.cell(row=rr, column=7, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*({GC_}=$B{rr})*{GG})+SUMPRODUCT(({GR}=$A{rr})*({GD}=$B{rr})*{GH})"))
            ws_t.cell(row=rr, column=8, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*({GC_}=$B{rr})*{GH})+SUMPRODUCT(({GR}=$A{rr})*({GD}=$B{rr})*{GG})"))
            ws_t.cell(row=rr, column=9, value=f"=G{rr}-H{rr}")
            ws_t.cell(row=rr, column=10, value=f"=D{rr}*3+E{rr}")
            ws_t.cell(row=rr, column=11, value=f"=J{rr}*1000000+I{rr}*1000+G{rr}*10-{ti}")
            # Pron stats (M..T = 13..20)
            ws_t.cell(row=rr, column=13, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*(({GC_}=$B{rr})+({GD}=$B{rr}))*({GE}<>\"\"))"))
            ws_t.cell(row=rr, column=14, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*({GC_}=$B{rr})*({GE}>{GF_}))+SUMPRODUCT(({GR}=$A{rr})*({GD}=$B{rr})*({GF_}>{GE}))"))
            ws_t.cell(row=rr, column=15, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*(({GC_}=$B{rr})+({GD}=$B{rr}))*({GE}={GF_})*({GE}<>\"\"))"))
            ws_t.cell(row=rr, column=16, value=f"=M{rr}-N{rr}-O{rr}")
            ws_t.cell(row=rr, column=17, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*({GC_}=$B{rr})*{GE})+SUMPRODUCT(({GR}=$A{rr})*({GD}=$B{rr})*{GF_})"))
            ws_t.cell(row=rr, column=18, value=(
                f"=SUMPRODUCT(({GR}=$A{rr})*({GC_}=$B{rr})*{GF_})+SUMPRODUCT(({GR}=$A{rr})*({GD}=$B{rr})*{GE})"))
            ws_t.cell(row=rr, column=19, value=f"=Q{rr}-R{rr}")
            ws_t.cell(row=rr, column=20, value=f"=N{rr}*3+O{rr}")
            ws_t.cell(row=rr, column=21, value=f"=T{rr}*1000000+S{rr}*1000+Q{rr}*10-{ti}")
            # Posicion index fijo (23) = ti+1
            ws_t.cell(row=rr, column=23, value=ti + 1)
            row += 1
        r3 = row - 1
        # Ranks (need full group range) — fill after knowing r0..r3
        for rr in range(r0, r3 + 1):
            ws_t.cell(row=rr, column=12, value=f"=1+COUNTIF($K${r0}:$K${r3},\">\"&K{rr})")
            ws_t.cell(row=rr, column=22, value=f"=1+COUNTIF($U${r0}:$U${r3},\">\"&U{rr})")
            ws_t.cell(row=rr, column=24, value=f"=INDEX($B${r0}:$B${r3},MATCH(W{rr},$L${r0}:$L${r3},0))")
            ws_t.cell(row=rr, column=25, value=f"=INDEX($B${r0}:$B${r3},MATCH(W{rr},$V${r0}:$V${r3},0))")
            ws_t.cell(row=rr, column=26, value=f"=IF(AND(X{rr}<>\"\",X{rr}=Y{rr}),1,0)")
            ws_t.cell(row=rr, column=27, value=f"=INDEX($J${r0}:$J${r3},MATCH(W{rr},$L${r0}:$L${r3},0))")
            ws_t.cell(row=rr, column=28, value=f"=INDEX($I${r0}:$I${r3},MATCH(W{rr},$L${r0}:$L${r3},0))")
            ws_t.cell(row=rr, column=29, value=f"=INDEX($G${r0}:$G${r3},MATCH(W{rr},$L${r0}:$L${r3},0))")
            ws_t.cell(row=rr, column=30, value=f"=INDEX($T${r0}:$T${r3},MATCH(W{rr},$V${r0}:$V${r3},0))")
            ws_t.cell(row=rr, column=31, value=f"=INDEX($S${r0}:$S${r3},MATCH(W{rr},$V${r0}:$V${r3},0))")
            ws_t.cell(row=rr, column=32, value=f"=INDEX($Q${r0}:$Q${r3},MATCH(W{rr},$V${r0}:$V${r3},0))")
            for c in ws_t[rr]:
                c.border = BORDER
        group_row_ranges[g] = (r0, r3)
    TEAM_LAST_ROW = row - 1

    pos_pts_row = TEAM_LAST_ROW + 2
    ws_t.cell(row=pos_pts_row, column=1, value='PUNTOS POSICION EXACTA (suma columna Z)').font = Font(bold=True)
    ws_t.cell(row=pos_pts_row, column=26, value=f"=SUM(Z{TEAM_FIRST_ROW}:Z{TEAM_LAST_ROW})").font = Font(bold=True)

    # ── Mejores Terceros (12 filas, 1 por grupo) ──
    mt_header_row = pos_pts_row + 3
    mt_headers = ['Grupo', 'EquipoReal3ro', 'PtsReal', 'DGReal', 'GFReal', 'OrdenReal', 'RankReal(1-12)', 'ClasificaReal',
                  'EquipoPron3ro', 'PtsPron', 'DGPron', 'GFPron', 'OrdenPron', 'RankPron(1-12)', 'ClasificaPron']
    for j, h in enumerate(mt_headers, start=1):
        ws_t.cell(row=mt_header_row, column=j, value=h)
        style_header(ws_t.cell(row=mt_header_row, column=j))
    MT_FIRST = mt_header_row + 1
    for gi, g in enumerate(GROUP_LETTERS):
        r0, r3 = group_row_ranges[g]
        pos3_row = r0 + 2  # tercer puesto = fila con W=3 dentro del bloque (0-indexed +2)
        rr = MT_FIRST + gi
        ws_t.cell(row=rr, column=1, value=g)
        ws_t.cell(row=rr, column=2, value=f"=X{pos3_row}")
        ws_t.cell(row=rr, column=3, value=f"=AA{pos3_row}")
        ws_t.cell(row=rr, column=4, value=f"=AB{pos3_row}")
        ws_t.cell(row=rr, column=5, value=f"=AC{pos3_row}")
        ws_t.cell(row=rr, column=6, value=f"=C{rr}*1000000+D{rr}*1000+E{rr}*10-{gi}")
        ws_t.cell(row=rr, column=9, value=f"=Y{pos3_row}")
        ws_t.cell(row=rr, column=10, value=f"=AD{pos3_row}")
        ws_t.cell(row=rr, column=11, value=f"=AE{pos3_row}")
        ws_t.cell(row=rr, column=12, value=f"=AF{pos3_row}")
        ws_t.cell(row=rr, column=13, value=f"=J{rr}*1000000+K{rr}*1000+L{rr}*10-{gi}")
        for c in ws_t[rr]:
            c.border = BORDER
    MT_LAST = MT_FIRST + 11
    for rr in range(MT_FIRST, MT_LAST + 1):
        ws_t.cell(row=rr, column=7, value=f"=1+COUNTIF($F${MT_FIRST}:$F${MT_LAST},\">\"&F{rr})")
        ws_t.cell(row=rr, column=8, value=f"=IF(G{rr}<=8,1,0)")
        ws_t.cell(row=rr, column=14, value=f"=1+COUNTIF($M${MT_FIRST}:$M${MT_LAST},\">\"&M{rr})")
        ws_t.cell(row=rr, column=15, value=f"=IF(N{rr}<=8,1,0)")

    # ── 8 mejores terceros (listas planas Real / Pron) ──
    best8_header_row = MT_LAST + 3
    ws_t.cell(row=best8_header_row, column=1, value='8 MEJORES TERCEROS').font = Font(bold=True)
    ws_t.cell(row=best8_header_row + 1, column=1, value='Puesto')
    ws_t.cell(row=best8_header_row + 1, column=2, value='Real')
    ws_t.cell(row=best8_header_row + 1, column=3, value='Pronostico')
    for c in ws_t[best8_header_row + 1][:3]:
        style_header(c)
    BEST8_FIRST = best8_header_row + 2
    for n in range(1, 9):
        rr = BEST8_FIRST + n - 1
        ws_t.cell(row=rr, column=1, value=n)
        ws_t.cell(row=rr, column=2, value=f"=IFERROR(INDEX($B${MT_FIRST}:$B${MT_LAST},MATCH({n},$G${MT_FIRST}:$G${MT_LAST},0)),\"\")")
        ws_t.cell(row=rr, column=3, value=f"=IFERROR(INDEX($I${MT_FIRST}:$I${MT_LAST},MATCH({n},$N${MT_FIRST}:$N${MT_LAST},0)),\"\")")
    BEST8_LAST = BEST8_FIRST + 7

    # ── Clasificados a 32 (Real / Pronostico), 32 filas cada uno ──
    cl_header_row = BEST8_LAST + 3
    ws_t.cell(row=cl_header_row, column=1, value='CLASIFICADOS A 32AVOS').font = Font(bold=True)
    ws_t.cell(row=cl_header_row + 1, column=1, value='#')
    ws_t.cell(row=cl_header_row + 1, column=2, value='Real')
    ws_t.cell(row=cl_header_row + 1, column=3, value='Pronostico')
    for c in ws_t[cl_header_row + 1][:3]:
        style_header(c)
    CL_FIRST = cl_header_row + 2
    idx = 0
    for g in GROUP_LETTERS:
        r0, r3 = group_row_ranges[g]
        pos1_row, pos2_row = r0, r0 + 1
        for pos_row in (pos1_row, pos2_row):
            rr = CL_FIRST + idx
            ws_t.cell(row=rr, column=1, value=idx + 1)
            ws_t.cell(row=rr, column=2, value=f"=X{pos_row}")
            ws_t.cell(row=rr, column=3, value=f"=Y{pos_row}")
            idx += 1
    for n in range(1, 9):
        rr = CL_FIRST + idx
        best8_row = BEST8_FIRST + n - 1
        ws_t.cell(row=rr, column=1, value=idx + 1)
        ws_t.cell(row=rr, column=2, value=f"=B{best8_row}")
        ws_t.cell(row=rr, column=3, value=f"=C{best8_row}")
        idx += 1
    CL_LAST = CL_FIRST + idx - 1  # deberia ser CL_FIRST+31

    cl_pts_row = CL_LAST + 2
    ws_t.cell(row=cl_pts_row, column=1, value='PUNTOS CLASIFICADO A 32 (2 pts/equipo)').font = Font(bold=True)
    cl_real_cells = [f"$B${rr}" for rr in range(CL_FIRST, CL_LAST + 1)]
    ws_t.cell(row=cl_pts_row, column=3,
              value=membership_formula(cl_real_cells, f"$C${CL_FIRST}:$C${CL_LAST}", 2)).font = Font(bold=True)

    autosize(ws_t, min_width=8, max_width=22)
    ws_t.freeze_panes = 'C2'

    POS_PTS_CELL = f"'Tabla Grupos'!Z{pos_pts_row}"
    CLASIFICADO_PTS_CELL = f"'Tabla Grupos'!C{cl_pts_row}"

    # ══════════════════════════ 5. BRACKET ══════════════════════════
    ws_b = wb.create_sheet('Bracket')
    bheaders = ['Ronda', '#', 'Pron. Local', 'Pron. Visitante', 'Pron. Ganador',
                'Real Local', 'Real Visitante', 'Real Ganador', 'Puntos Llave']
    ws_b.append(bheaders)
    for c in ws_b[1]:
        style_header(c)

    round_rows = {}
    row = 2
    for rnd in ROUND_ORDER:
        ms = sorted([m for m in bracket if m['round'] == rnd], key=lambda m: m['position'])
        r0 = row
        for m in ms:
            ws_b.append([ROUND_LABELS[rnd], m['position'], None, None, None, None, None, None, None])
            rr = ws_b.max_row
            pts = ROUND_PAIR_PTS[rnd]
            f = (f"=IF(OR(C{rr}=\"\",D{rr}=\"\",F{rr}=\"\",G{rr}=\"\"),0,"
                 f"IF(OR(AND(C{rr}=F{rr},D{rr}=G{rr}),AND(C{rr}=G{rr},D{rr}=F{rr})),{pts},0))")
            ws_b.cell(row=rr, column=9, value=f)
            for col in (3, 4, 5, 6, 7, 8):
                ws_b.cell(row=rr, column=col).fill = INPUT_FILL
            ws_b.cell(row=rr, column=9).fill = CALC_FILL
            for c in ws_b[rr]:
                c.border = BORDER
            row += 1
        round_rows[rnd] = (r0, row - 1)
    BRACKET_LAST_DATA_ROW = row - 1

    # Listas desplegables de equipo (evita variaciones de escritura como con el goleador)
    n_teams = sum(len(t) for t in groups_teams.values())
    dv_team = DataValidation(type='list', formula1=f"=Equipos!$B$2:$B${1 + n_teams}", allow_blank=True)
    ws_b.add_data_validation(dv_team)
    for col in ('C', 'D', 'F', 'G'):
        dv_team.add(f"{col}2:{col}{BRACKET_LAST_DATA_ROW}")
    dv_winner = DataValidation(type='list', formula1=f"=Equipos!$B$2:$B${1 + n_teams}", allow_blank=True)
    ws_b.add_data_validation(dv_winner)
    for col in ('E', 'H'):
        dv_winner.add(f"{col}2:{col}{BRACKET_LAST_DATA_ROW}")

    r32_0, r32_1 = round_rows['r32']
    r16_0, r16_1 = round_rows['r16']
    qf_0, qf_1 = round_rows['qf']
    sf_0, sf_1 = round_rows['sf']
    th_0, th_1 = round_rows['third']
    fi_0, fi_1 = round_rows['final']

    def real_cells_fg(r0, r1):
        return [f"F{rr}" for rr in range(r0, r1 + 1)] + [f"G{rr}" for rr in range(r0, r1 + 1)]

    r16_team_f = membership_formula(real_cells_fg(r16_0, r16_1), f"C{r16_0}:D{r16_1}", 2)
    qf_team_f = membership_formula(real_cells_fg(qf_0, qf_1), f"C{qf_0}:D{qf_1}", 4)
    sf_team_f = membership_formula(real_cells_fg(sf_0, sf_1), f"C{sf_0}:D{sf_1}", 4)
    final_team_f = membership_formula(real_cells_fg(fi_0, fi_1), f"C{fi_0}:D{fi_1}", 8)
    third_team_f = membership_formula(real_cells_fg(th_0, th_1), f"C{th_0}:D{th_1}", 8)
    # Combina Final + Tercer puesto como formula unica (cada set se cuenta por separado, no se mezclan)
    finalfour_team_f = '=' + final_team_f[1:] + '+' + third_team_f[1:]

    summary_row = row + 2
    ws_b.cell(row=summary_row, column=1, value='RESUMEN BRACKET (equipo clasificado + llave)').font = Font(bold=True)
    labels_formulas = [
        ('32avos — Llave (3 pts c/u)', f"=SUM(I{r32_0}:I{r32_1})"),
        ('Octavos — Equipo (2 pts c/u)', r16_team_f),
        ('Octavos — Llave (4 pts c/u)', f"=SUM(I{r16_0}:I{r16_1})"),
        ('Cuartos — Equipo (4 pts c/u)', qf_team_f),
        ('Cuartos — Llave (8 pts c/u)', f"=SUM(I{qf_0}:I{qf_1})"),
        ('Semifinal — Equipo (4 pts c/u)', sf_team_f),
        ('Semifinal — Llave (8 pts c/u)', f"=SUM(I{sf_0}:I{sf_1})"),
        ('Final/3er — Equipo (8 pts c/u)', finalfour_team_f),
        ('3er puesto — Llave (8 pts)', f"=I{th_0}"),
        ('Final — Llave (12 pts)', f"=I{fi_0}"),
        ('Bonus Campeon (15 pts)', f"=IF(AND(E{fi_0}<>\"\",H{fi_0}<>\"\",E{fi_0}=H{fi_0}),15,0)"),
        ('Bonus Subcampeon (10 pts)',
         f"=IF(AND(C{fi_0}<>\"\",D{fi_0}<>\"\",F{fi_0}<>\"\",G{fi_0}<>\"\",E{fi_0}<>\"\",H{fi_0}<>\"\"),"
         f"IF(IF(E{fi_0}=C{fi_0},D{fi_0},C{fi_0})=IF(H{fi_0}=F{fi_0},G{fi_0},F{fi_0}),10,0),0)"),
        ('Bonus 3er puesto (8 pts)', f"=IF(AND(E{th_0}<>\"\",H{th_0}<>\"\",E{th_0}=H{th_0}),8,0)"),
        ('Bonus 4to puesto (6 pts)',
         f"=IF(AND(C{th_0}<>\"\",D{th_0}<>\"\",F{th_0}<>\"\",G{th_0}<>\"\",E{th_0}<>\"\",H{th_0}<>\"\"),"
         f"IF(IF(E{th_0}=C{th_0},D{th_0},C{th_0})=IF(H{th_0}=F{th_0},G{th_0},F{th_0}),6,0),0)"),
    ]
    r = summary_row + 1
    bracket_term_cells = []
    for label, formula in labels_formulas:
        ws_b.cell(row=r, column=1, value=label)
        cell = ws_b.cell(row=r, column=3, value=formula)
        cell.fill = CALC_FILL
        bracket_term_cells.append(f"'Bracket'!C{r}")
        r += 1
    bracket_total_row = r + 1
    ws_b.cell(row=bracket_total_row, column=1, value='TOTAL BRACKET (equipo+llave+bonus)').font = Font(bold=True)
    total_cell = ws_b.cell(row=bracket_total_row, column=3, value='=' + '+'.join(f"C{summary_row + 1 + i}" for i in range(len(labels_formulas))))
    total_cell.font = Font(bold=True)
    total_cell.fill = GOLD_FILL

    autosize(ws_b, min_width=10, max_width=30)
    ws_b.freeze_panes = 'A2'
    ws_b.column_dimensions['A'].width = 30

    BRACKET_TOTAL_CELL = f"'Bracket'!C{bracket_total_row}"

    # ══════════════════════════ 6. GOLEADOR ══════════════════════════
    ws_s = wb.create_sheet('Goleador')
    ws_s.append(['Pron. Goleador', 'Real Goleador 1', 'Real Goleador 2', 'Real Goleador 3', 'Puntos'])
    for c in ws_s[1]:
        style_header(c)
    ws_s.append([None, None, None, None, None])
    f = ('=IF(TRIM(A2)="",0,IF(OR(EXACT(UPPER(TRIM(A2)),UPPER(TRIM(B2))),'
         'EXACT(UPPER(TRIM(A2)),UPPER(TRIM(C2))),EXACT(UPPER(TRIM(A2)),UPPER(TRIM(D2)))),10,0))')
    ws_s.cell(row=2, column=5, value=f)
    for col in (1, 2, 3, 4):
        ws_s.cell(row=2, column=col).fill = INPUT_FILL
    ws_s.cell(row=2, column=5).fill = CALC_FILL
    for c in ws_s[2]:
        c.border = BORDER
    autosize(ws_s, min_width=16, max_width=30)

    GOLEADOR_PTS_CELL = "Goleador!E2"

    # ══════════════════════════ 7. RESUMEN PUNTAJE ══════════════════════════
    ws_r = wb.create_sheet('Resumen Puntaje')
    ws_r.append(['Fase', 'Puntos'])
    for c in ws_r[1]:
        style_header(c)
    rows_resumen = [
        ('Grupos — marcador (exacto + resultado)', f"=Grupos!I{GROUPS_LAST_ROW + 1}"),
        ('Grupos — posicion exacta en tabla', f"={POS_PTS_CELL}"),
        ('32avos — clasificado a dieciseisavos', f"={CLASIFICADO_PTS_CELL}"),
        ('32avos — llave del bracket', f"='Bracket'!C{summary_row + 1}"),
        ('16avos (Octavos) — equipo + llave', f"='Bracket'!C{summary_row + 2}+'Bracket'!C{summary_row + 3}"),
        ('Cuartos — equipo + llave', f"='Bracket'!C{summary_row + 4}+'Bracket'!C{summary_row + 5}"),
        ('Semifinal — equipo + llave', f"='Bracket'!C{summary_row + 6}+'Bracket'!C{summary_row + 7}"),
        ('Final / 3er puesto — equipo + llaves + bonus',
         f"='Bracket'!C{summary_row + 8}+'Bracket'!C{summary_row + 9}+'Bracket'!C{summary_row + 10}"
         f"+'Bracket'!C{summary_row + 11}+'Bracket'!C{summary_row + 12}+'Bracket'!C{summary_row + 13}+'Bracket'!C{summary_row + 14}"),
        ('Goleador', f"={GOLEADOR_PTS_CELL}"),
    ]
    for label, formula in rows_resumen:
        ws_r.append([label, formula])
        ws_r.cell(row=ws_r.max_row, column=2).fill = CALC_FILL
        for c in ws_r[ws_r.max_row]:
            c.border = BORDER
    total_row = ws_r.max_row + 1
    ws_r.cell(row=total_row, column=1, value='TOTAL').font = Font(bold=True, size=12)
    ws_r.cell(row=total_row, column=2, value=f"=SUM(B2:B{total_row - 1})").font = Font(bold=True, size=12)
    ws_r.cell(row=total_row, column=2).fill = GOLD_FILL
    ws_r.cell(row=total_row, column=1).fill = GOLD_FILL
    autosize(ws_r, min_width=20, max_width=55)
    ws_r.column_dimensions['A'].width = 48

    wb.save(output_path)
    print('OK ->', output_path)
    print('debug rows:', dict(
        GROUPS_LAST_ROW=GROUPS_LAST_ROW, TEAM_LAST_ROW=TEAM_LAST_ROW, pos_pts_row=pos_pts_row,
        MT_FIRST=MT_FIRST, MT_LAST=MT_LAST, BEST8_FIRST=BEST8_FIRST, BEST8_LAST=BEST8_LAST,
        CL_FIRST=CL_FIRST, CL_LAST=CL_LAST, cl_pts_row=cl_pts_row, bracket_total_row=bracket_total_row,
    ))


if __name__ == '__main__':
    main()
