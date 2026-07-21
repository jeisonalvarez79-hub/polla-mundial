"""
Convierte el backup JSON de pronósticos (scripts/generate-backup.mjs) en un
Excel entendible: una fila por participante, columnas por cada partido de
grupos y cada llave de bracket, con su pronóstico.
Ejecutar: python scripts/build-backup-excel.py <input.json> <output.xlsx>
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='2F6B4D')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=9)
GROUP_FILL_A = PatternFill('solid', fgColor='EEF2EA')
GROUP_FILL_B = PatternFill('solid', fgColor='FFFFFF')
NAME_FONT = Font(bold=True, size=10)
THIN = Side(style='thin', color='CCD4C3')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROUND_LABELS = {'r32': '32avos', 'r16': '16avos', 'qf': 'Cuartos', 'sf': 'Semifinal', 'third': '3er Puesto', 'final': 'Final'}
ROUND_ORDER = ['r32', 'r16', 'qf', 'sf', 'third', 'final']


def autosize(ws, min_width=10, max_width=42):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def main():
    input_path, output_path = sys.argv[1], sys.argv[2]
    with open(input_path, encoding='utf-8') as f:
        data = json.load(f)

    poll_name_by_id = {p['id']: p['name'] for p in data['pollas']}
    participants = sorted(data['participants'], key=lambda p: (poll_name_by_id.get(p['pollaId'], ''), p['name']))
    part_ids = {p['id'] for p in participants}

    pred_by_participant = {}
    for p in data['predictions']:
        pred_by_participant.setdefault(p['participantId'], {})[p['matchId']] = p

    bpred_by_participant = {}
    for p in data['bracketPredictions']:
        bpred_by_participant.setdefault(p['participantId'], {})[p['bracketMatchId']] = p

    scorer_by_participant = {s['participantId']: s for s in data['scorerPredictions']}

    group_matches = sorted([m for m in data['matches'] if m['phase'] == 'groups'], key=lambda m: (m['group'] or '', m['id']))
    bracket_matches = {m['id']: m for m in data['bracketMatches']}

    wb = Workbook()

    # ── Sheet 1: Participantes ──
    ws0 = wb.active
    ws0.title = 'Participantes'
    ws0.append(['Polla', 'Participante', 'ID interno', 'Fecha de registro'])
    for c in ws0[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for p in participants:
        ws0.append([poll_name_by_id.get(p['pollaId'], '?'), p['name'], p['id'], (p['createdAt'] or '')[:10]])
    autosize(ws0)
    ws0.freeze_panes = 'A2'

    # ── Sheet 2: Grupos (grid) ──
    ws1 = wb.create_sheet('Pronosticos Grupos')
    header1 = ['Polla', 'Participante']
    for m in group_matches:
        header1.append(f"{m['group']}: {m['homeTeam']} vs {m['awayTeam']}")
    ws1.append(header1)
    for c in ws1[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical='center')

    row_i = 2
    last_group_shade = None
    for p in participants:
        row = [poll_name_by_id.get(p['pollaId'], '?'), p['name']]
        preds = pred_by_participant.get(p['id'], {})
        for m in group_matches:
            pred = preds.get(m['id'])
            if pred and pred['homeScore'] is not None and pred['awayScore'] is not None:
                row.append(f"{pred['homeScore']}-{pred['awayScore']}")
            else:
                row.append('')
        ws1.append(row)
        for c in ws1[row_i]:
            c.alignment = Alignment(horizontal='center')
            c.border = BORDER
        ws1.cell(row=row_i, column=2).font = NAME_FONT
        ws1.cell(row=row_i, column=2).alignment = Alignment(horizontal='left')
        ws1.cell(row=row_i, column=1).alignment = Alignment(horizontal='left')
        row_i += 1
    ws1.freeze_panes = 'C2'
    ws1.row_dimensions[1].height = 60
    for col_cells in ws1.iter_cols(min_col=1, max_col=2):
        col_letter = get_column_letter(col_cells[0].column)
        ws1.column_dimensions[col_letter].width = 20
    for col_idx in range(3, 3 + len(group_matches)):
        ws1.column_dimensions[get_column_letter(col_idx)].width = 13

    # ── Sheet 3: Bracket (grid) ──
    ws2 = wb.create_sheet('Pronosticos Bracket')
    bracket_ids_ordered = []
    for rnd in ROUND_ORDER:
        ids = sorted([mid for mid, m in bracket_matches.items() if m['round'] == rnd],
                     key=lambda mid: bracket_matches[mid].get('position') or 0)
        bracket_ids_ordered.extend(ids)

    header2 = ['Polla', 'Participante']
    for mid in bracket_ids_ordered:
        m = bracket_matches[mid]
        if m.get('homeTeam') or m.get('awayTeam'):
            matchup = f"{m.get('homeTeam') or '?'} vs {m.get('awayTeam') or '?'}"
        else:
            matchup = m.get('label') or f"#{m.get('position') or ''}"
        header2.append(f"{ROUND_LABELS.get(m['round'], m['round'])} #{m.get('position') or ''}: {matchup}")
    ws2.append(header2)
    for c in ws2[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(wrap_text=True, vertical='center')

    row_i = 2
    for p in participants:
        row = [poll_name_by_id.get(p['pollaId'], '?'), p['name']]
        bpreds = bpred_by_participant.get(p['id'], {})
        for mid in bracket_ids_ordered:
            bp = bpreds.get(mid)
            row.append(bp['predictedWinner'] if bp and bp.get('predictedWinner') else '')
        ws2.append(row)
        for c in ws2[row_i]:
            c.alignment = Alignment(horizontal='center')
            c.border = BORDER
        ws2.cell(row=row_i, column=2).font = NAME_FONT
        ws2.cell(row=row_i, column=2).alignment = Alignment(horizontal='left')
        ws2.cell(row=row_i, column=1).alignment = Alignment(horizontal='left')
        row_i += 1
    ws2.freeze_panes = 'C2'
    ws2.row_dimensions[1].height = 60
    for col_cells in ws2.iter_cols(min_col=1, max_col=2):
        col_letter = get_column_letter(col_cells[0].column)
        ws2.column_dimensions[col_letter].width = 20
    for col_idx in range(3, 3 + len(bracket_ids_ordered)):
        ws2.column_dimensions[get_column_letter(col_idx)].width = 16

    # ── Sheet 4: Goleador ──
    ws3 = wb.create_sheet('Pronostico Goleador')
    ws3.append(['Polla', 'Participante', 'Goleador pronosticado (1o)', 'Otros (2o/3o)'])
    for c in ws3[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
    for p in participants:
        s = scorer_by_participant.get(p['id'])
        scorers = (s.get('scorers') or []) if s else []
        first = scorers[0] if scorers else ''
        rest = ', '.join([x for x in scorers[1:] if x])
        ws3.append([poll_name_by_id.get(p['pollaId'], '?'), p['name'], first, rest])
    autosize(ws3)
    ws3.freeze_panes = 'A2'

    # ── Sheet 5: Notas del backup ──
    ws4 = wb.create_sheet('Notas')
    notes = [
        ['Backup de pronosticos - Polla Mundial 2026'],
        [f"Generado: {data['exportedAt']}"],
        [f"Participantes: {data['counts']['participants']}"],
        [f"Partidos de grupos: {data['counts']['matches']}"],
        [f"Llaves de bracket: {data['counts']['bracketMatches']}"],
        [f"Pronosticos de grupos guardados: {data['counts']['predictions']}"],
        [f"Pronosticos de bracket guardados: {data['counts']['bracketPredictions']}"],
        [''],
        ['No incluye el PIN de los participantes (dato de login, fuera del alcance de este backup de pronosticos).'],
        ['El archivo JSON del mismo backup contiene los datos crudos completos para restauracion si se necesitara.'],
    ]
    for r in notes:
        ws4.append(r)
    ws4.column_dimensions['A'].width = 90

    wb.save(output_path)
    print('OK ->', output_path)


if __name__ == '__main__':
    main()
