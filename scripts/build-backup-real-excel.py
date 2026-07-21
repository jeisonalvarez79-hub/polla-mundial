"""
Convierte el backup JSON de resultados REALES (scripts/generate-backup-real.mjs)
en un Excel entendible: partidos de grupos, tabla de posiciones, clasificados,
bracket real y goleador.
Ejecutar: python scripts/build-backup-real-excel.py <input.json> <output.xlsx>
"""
import json
import sys

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill('solid', fgColor='2F6B4D')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=10)
GOLD_FILL = PatternFill('solid', fgColor='FBF1DE')
NAME_FONT = Font(bold=True, size=10)
THIN = Side(style='thin', color='CCD4C3')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ROUND_LABELS = {'r32': '32avos', 'r16': '16avos', 'qf': 'Cuartos', 'sf': 'Semifinal', 'third': '3er Puesto', 'final': 'Final'}
ROUND_ORDER = ['r32', 'r16', 'qf', 'sf', 'third', 'final']


def header_row(ws, values):
    ws.append(values)
    for c in ws[ws.max_row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal='center', vertical='center')


def autosize(ws, min_width=10, max_width=42):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        col_letter = get_column_letter(col_cells[0].column)
        ws.column_dimensions[col_letter].width = max(min_width, min(max_width, length + 2))


def main():
    input_path, output_path = sys.argv[1], sys.argv[2]
    with open(input_path, encoding='utf-8') as f:
        data = json.load(f)

    wb = Workbook()

    # ── Sheet 1: Partidos de Grupos (real) ──
    ws0 = wb.active
    ws0.title = 'Partidos Grupos (Real)'
    header_row(ws0, ['Grupo', 'Partido', 'Marcador Real', 'Status'])
    group_matches = sorted([m for m in data['matches'] if m['phase'] == 'groups'], key=lambda m: (m['group'] or '', m['id']))
    for m in group_matches:
        score = f"{m['homeScore']}-{m['awayScore']}" if m['homeScore'] is not None and m['awayScore'] is not None else '—'
        ws0.append([m['group'], f"{m['homeTeam']} vs {m['awayTeam']}", score, m['status']])
    for row in ws0.iter_rows(min_row=2):
        for c in row:
            c.border = BORDER
        row[2].alignment = Alignment(horizontal='center')
    autosize(ws0)
    ws0.freeze_panes = 'A2'

    # ── Sheet 2: Tabla de Grupos (real) ──
    ws1 = wb.create_sheet('Tabla de Grupos (Real)')
    header_row(ws1, ['Grupo', 'Pos', 'Equipo', 'J', 'G', 'E', 'P', 'GF', 'GC', 'DG', 'Pts', 'Clasifica'])
    for group in sorted(data['groupStandings'].keys()):
        info = data['groupStandings'][group]
        for i, t in enumerate(info['standings']):
            clasifica = '1o/2o directo' if i < 2 else ('Posible mejor 3o' if i == 2 else '')
            ws1.append([group, i + 1, t['name'], t['J'], t['G'], t['E'], t['P'], t['GF'], t['GC'], t['DG'], t['Pts'], clasifica])
    for row in ws1.iter_rows(min_row=2):
        for c in row:
            c.border = BORDER
        for c in row[1:11]:
            c.alignment = Alignment(horizontal='center')
    autosize(ws1)
    ws1.freeze_panes = 'A2'

    # ── Sheet 3: Clasificados a 32avos (real) ──
    ws2 = wb.create_sheet('Clasificados 32avos (Real)')
    header_row(ws2, ['Slot', 'Equipo', 'Tipo'])
    if data.get('r32Qualifiers'):
        items = list(data['r32Qualifiers'].items())
        items.sort(key=lambda kv: (0 if kv[0][0] in '12' else 1, kv[0]))
        for slot, team in items:
            tipo = 'Mejor 3ero' if slot.startswith('t') else ('1o de grupo' if slot.startswith('1') else '2o de grupo')
            ws2.append([slot, team, tipo])
        for row in ws2.iter_rows(min_row=2):
            for c in row:
                c.border = BORDER
    else:
        ws2.append(['(fase de grupos aun no finaliza por completo)', '', ''])
    autosize(ws2)
    ws2.freeze_panes = 'A2'

    # ── Sheet 4: Bracket Real ──
    ws3 = wb.create_sheet('Bracket (Real)')
    header_row(ws3, ['Ronda', '#', 'Partido', 'Marcador', 'Ganador', 'Status'])
    bracket_sorted = []
    for rnd in ROUND_ORDER:
        ms = [m for m in data['bracketMatches'] if m['round'] == rnd]
        ms.sort(key=lambda m: m.get('position') or 0)
        bracket_sorted.extend(ms)
    for m in bracket_sorted:
        score = f"{m['homeScore']}-{m['awayScore']}" if m['homeScore'] is not None and m['awayScore'] is not None else '—'
        ws3.append([ROUND_LABELS.get(m['round'], m['round']), m.get('position') or '', f"{m['homeTeam'] or '?'} vs {m['awayTeam'] or '?'}",
                    score, m.get('winner') or '', m['status']])
    for row in ws3.iter_rows(min_row=2):
        for c in row:
            c.border = BORDER
        row[1].alignment = Alignment(horizontal='center')
        row[3].alignment = Alignment(horizontal='center')
        if row[4].value:
            row[4].font = Font(bold=True)
            row[4].fill = GOLD_FILL
    autosize(ws3)
    ws3.freeze_panes = 'A2'

    # ── Sheet 5: Goleador (real) ──
    ws4 = wb.create_sheet('Goleador (Real)')
    header_row(ws4, ['Top Goleadores Oficiales'])
    for s in data['topScorers']:
        if s:
            ws4.append([s])
    autosize(ws4)

    # ── Sheet 6: Notas ──
    ws5 = wb.create_sheet('Notas')
    notes = [
        ['Backup de RESULTADOS REALES - Polla Mundial 2026'],
        [f"Generado: {data['exportedAt']}"],
        [f"Partidos de grupos: {data['counts']['matches']} (finalizados: {data['counts']['matchesFinished']})"],
        [f"Llaves de bracket: {data['counts']['bracketMatches']} (finalizadas: {data['counts']['bracketMatchesFinished']})"],
        [''],
        ['Este backup contiene los RESULTADOS REALES del torneo (marcadores, tabla de grupos, clasificados y bracket),'],
        ['no los pronosticos de los participantes (ver backup-polla-pronosticos-*.xlsx para eso).'],
        ['El archivo JSON del mismo backup contiene los datos crudos completos para restauracion si se necesitara.'],
    ]
    for r in notes:
        ws5.append(r)
    ws5.column_dimensions['A'].width = 100

    wb.save(output_path)
    print('OK ->', output_path)


if __name__ == '__main__':
    main()
